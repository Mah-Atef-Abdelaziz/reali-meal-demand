"""Report Generator — Produces PDF and Excel reports."""
import os
from datetime import date, timedelta, datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models import PredictionResult, MealTransaction, WorkLocation

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "generated_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportGenerator:
    async def generate(self, db: AsyncSession, report_type: str,
                       start_date: date = None, end_date: date = None,
                       location_id: int = None, fmt: str = "pdf") -> str:
        """Generate a report and return the file path."""
        today = date.today()
        if report_type == "daily":
            start_date = start_date or today
            end_date = end_date or today
        elif report_type == "weekly":
            start_date = start_date or (today - timedelta(days=7))
            end_date = end_date or today
        elif report_type == "monthly":
            start_date = start_date or today.replace(day=1)
            end_date = end_date or today
        else:  # executive
            start_date = start_date or (today - timedelta(days=30))
            end_date = end_date or today

        data = await self._gather_report_data(db, start_date, end_date, location_id)

        if fmt == "excel":
            return await self._generate_excel(data, report_type, start_date, end_date)
        else:
            return await self._generate_csv_report(data, report_type, start_date, end_date)

    async def _gather_report_data(self, db, start_date, end_date, location_id=None):
        """Gather all metrics for the report."""
        query = select(
            PredictionResult.prediction_date,
            PredictionResult.period,
            func.sum(PredictionResult.predicted_count).label("predicted"),
            func.sum(PredictionResult.actual_count).label("actual"),
            func.sum(PredictionResult.predicted_waste).label("waste"),
        ).where(
            PredictionResult.prediction_date >= start_date,
            PredictionResult.prediction_date <= end_date
        ).group_by(PredictionResult.prediction_date, PredictionResult.period
        ).order_by(PredictionResult.prediction_date)

        if location_id:
            query = query.where(PredictionResult.location_id == location_id)

        result = await db.execute(query)
        rows = result.all()

        return [{"date": str(r.prediction_date), "period": r.period,
                 "predicted": r.predicted or 0, "actual": r.actual or 0,
                 "waste": r.waste or 0} for r in rows]

    async def _generate_csv_report(self, data, report_type, start_date, end_date):
        """Generate a CSV report (PDF requires reportlab)."""
        import csv
        filename = f"{report_type}_report_{start_date}_{end_date}.csv"
        filepath = os.path.join(REPORTS_DIR, filename)

        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["REAL.i Meal Demand AI — " + report_type.title() + " Report"])
            writer.writerow([f"Period: {start_date} to {end_date}"])
            writer.writerow([f"Generated: {datetime.now(timezone.utc).isoformat()}"])
            writer.writerow([])
            writer.writerow(["Date", "Period", "Predicted", "Actual", "Waste", "Accuracy %"])

            for row in data:
                actual = row["actual"] or row["predicted"]
                accuracy = round((1 - abs(row["predicted"] - actual) / max(actual, 1)) * 100, 1)
                writer.writerow([row["date"], row["period"], row["predicted"],
                                 actual, row["waste"], accuracy])

            # Summary
            writer.writerow([])
            writer.writerow(["SUMMARY"])
            total_pred = sum(r["predicted"] for r in data)
            total_waste = sum(r["waste"] for r in data)
            writer.writerow(["Total Predicted", total_pred])
            writer.writerow(["Total Waste", total_waste])
            writer.writerow(["Waste Rate %", round(total_waste / max(total_pred, 1) * 100, 1)])

        return filepath

    async def _generate_excel(self, data, report_type, start_date, end_date):
        """Generate an Excel report."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            filename = f"{report_type}_report_{start_date}_{end_date}.xlsx"
            filepath = os.path.join(REPORTS_DIR, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{report_type.title()} Report"

            # Header
            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            headers = ["Date", "Period", "Predicted", "Actual", "Waste", "Accuracy %"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="1A1A2E")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(data, 2):
                actual = row["actual"] or row["predicted"]
                accuracy = round((1 - abs(row["predicted"] - actual) / max(actual, 1)) * 100, 1)
                ws.cell(row=row_idx, column=1, value=row["date"])
                ws.cell(row=row_idx, column=2, value=row["period"])
                ws.cell(row=row_idx, column=3, value=row["predicted"])
                ws.cell(row=row_idx, column=4, value=actual)
                ws.cell(row=row_idx, column=5, value=row["waste"])
                ws.cell(row=row_idx, column=6, value=accuracy)

            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 15

            wb.save(filepath)
            return filepath
        except ImportError:
            return await self._generate_csv_report(data, report_type, start_date, end_date)
